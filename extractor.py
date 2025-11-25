"""
Submittal Extractor - PyMuPDF Version
Optimized for performance with caching and pre-compiled patterns
Handles multi-line subsection headings
"""

try:
    import pymupdf as fitz  # PyMuPDF 1.24+
except ImportError:
    import fitz  # PyMuPDF older versions

import re
from typing import List, Dict, Tuple, Optional
import logging
import os
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class SubmittalExtractor:
    """
    Optimized extractor for CSI-formatted construction specification books

    PERFORMANCE IMPROVEMENTS:
    - Extracts full PDF text once at initialization
    - Pre-compiles all regex patterns
    - Uses string searching instead of repeated PDF parsing
    - Efficient section boundary detection
    - Handles multi-line subsection headings
    """

    # Submittal section keywords
    SUBMITTAL_KEYWORDS = [
        "ACTION SUBMITTALS",
        "INFORMATIONAL SUBMITTALS",
        "CLOSEOUT SUBMITTALS",
        "SHOP DRAWING SUBMITTALS",
        "AS-BUILT SUBMITTALS",
        "As-Built Submittals",
        "QUALITY ASSURANCE SUBMITTALS",
        "SUBMITTAL REQUIREMENTS",
        "SUBMITTAL SCHEDULE",
        "FORM OF SUBMITTALS",
        "RECORD DOCUMENT SUBMITTALS",
        "SUBMITTALS"
    ]

    def __init__(self, pdf_path: str, use_ocr: bool = False):
        """
        Initialize the extractor

        Args:
            pdf_path: Path to the PDF spec book
            use_ocr: Whether to use OCR for scanned PDFs
        """
        self.pdf_path = pdf_path
        self.use_ocr = use_ocr
        self.doc = None
        self.full_text = ""  # Cache for full PDF text
        self.toc = []
        self.sections_with_submittals = []
        self.sections_data = []

        # PRE-COMPILE ALL REGEX PATTERNS
        self._compile_patterns()

    def _compile_patterns(self):
        """Pre-compile all regex patterns for performance"""
        # Section header pattern - matches SECTION followed by numbers (with or without spaces)
        # Examples: "SECTION 011100" or "SECTION 01 11 00"
        self.section_header_pattern = re.compile(
            r'SECTION\s+(\d{2}\s+\d{2}\s+\d{2}|\d{5,6}(?:\.\d+)?)',
            re.IGNORECASE
        )

        # TOC line pattern - matches section numbers with or without spaces
        # Examples: "011100 Summary" or "01 11 00 Summary" or "01300 SUBMITTALS"
        self.toc_line_pattern = re.compile(
            r'^(\d{2}\s+\d{2}\s+\d{2}|\d{5,6}(?:\.\d+)?)\s+(.+?)(?:\s+\d+\s*)?$'
        )

        # Subsection number only (for multi-line detection)
        self.subsection_num_only_pattern = re.compile(r'^(\d+\.\d+)$')

        # Footer patterns
        self.footer_patterns = [
            re.compile(r'\d{5,}-\d+'),
            re.compile(r'\d{1,2}/\d{1,2}/\d{4}'),
            re.compile(r'©\s*\d{4}'),
            re.compile(r'Page \d+ of \d+'),
            re.compile(r'All Rights Reserved'),
            re.compile(r'^\d{5}\s*-\s*\d+$'),
            re.compile(r'CDM Smith'),
        ]

        # Hierarchy patterns based on user specification
        # Level 1: PART 1 - Section Heading
        # Level 2: 1.01, 1.02, 1.1, 1.04 (subsection headings)
        # Level 3: A., B., C. (paragraphs)
        # Level 4: 1., 2., 3. (subparagraphs)
        # Level 5: a., b., c. (subparagraphs)
        # Level 6: 1), 2), 3) (subparagraphs)
        # Level 7: a), b), c) (subparagraphs)
        # Level 8: (a), (b), (c) (subparagraphs)
        # Level 9: (1), (2), (3) (subparagraphs)
        # Level 10: i., ii., iii. (subparagraphs)

        self.hierarchy_patterns = [
            re.compile(r'^PART\s+\d+\s*[-–—]?\s*(.*)$', re.IGNORECASE),  # Level 1: PART 1
            re.compile(r'^(\d+\.\d+)\s*[-–]?\s*(.+?)$'),                  # Level 2: 1.01, 1.3, 1.3 - TEXT
            re.compile(r'^([A-Z])\.\s*(.+?)$'),                           # Level 3: A., A.Text (requires dot)
            re.compile(r'^(\d{1,2})\.\s+(.+?)$'),                         # Level 4: 1., 2., ..., 99. (1-2 digits only)
            re.compile(r'^([a-z])\.\s+(.+?)$'),                           # Level 5: a., b. (requires dot + space)
            re.compile(r'^(\d+)\)\s+(.+?)$'),                             # Level 6: 1), 2)
            re.compile(r'^([a-z])\)\s+(.+?)$'),                           # Level 7: a), b)
            re.compile(r'^\(([a-z])\)\s+(.+?)$'),                         # Level 8: (a), (b)
            re.compile(r'^\((\d+)\)\s+(.+?)$'),                           # Level 9: (1), (2)
            re.compile(r'^([ivxlcdm]+)\.\s+(.+?)$'),                      # Level 10: i., ii.
        ]

    def extract(self, template_path: str = None) -> Dict:
        """
        Main extraction method - OPTIMIZED

        Args:
            template_path: Path to SubmittalLog.xlsx template

        Returns:
            Dictionary containing workbooks for sections and log
        """
        logger.info(f"Starting optimized extraction from: {self.pdf_path}")

        try:
            # Open PDF ONCE
            self.doc = fitz.open(self.pdf_path)
            logger.info(f"Opened PDF with {len(self.doc)} pages")

            # Extract ALL text ONCE and cache it
            logger.info("Extracting full PDF text (this happens once)...")
            self.full_text = self._extract_full_pdf_text()
            logger.info(f"Extracted {len(self.full_text)} characters")

            # Parse TOC from first 100 pages
            logger.info("Parsing Table of Contents from first 100 pages...")
            self.toc = self._extract_toc_from_first_100_pages()
            logger.info(f"Found {len(self.toc)} sections in TOC")

            # Fallback: If no TOC or TOC has no technical specs, scan PDF directly
            if len(self.toc) == 0:
                logger.warning("No sections found in TOC. Scanning PDF directly for SECTION headers...")
                self.toc = self._scan_pdf_for_sections()
                logger.info(f"Found {len(self.toc)} sections by scanning PDF")

            # Process sections
            logger.info("Processing sections...")
            for section_num_display, section_num_search, section_name in self.toc:
                self._process_section(section_num_display, section_num_search, section_name)

            logger.info(f"Extraction complete: {len(self.sections_with_submittals)} sections with submittals")

            return {
                "sections": self._format_sections_output(),
                "log": self._format_log_output(template_path)
            }

        except Exception as e:
            logger.error(f"Error during extraction: {str(e)}", exc_info=True)
            raise

        finally:
            if self.doc:
                self.doc.close()

    def _extract_full_pdf_text(self) -> str:
        """
        Extract all text from PDF in one go and cache it
        """
        full_text = []
        for page_num in range(len(self.doc)):
            page = self.doc[page_num]
            text = page.get_text("text", sort=True)

            # OCR if needed
            if self.use_ocr and len(text.strip()) < 100:
                try:
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                    img_data = pix.tobytes("png")

                    import pytesseract
                    from PIL import Image
                    import io

                    img = Image.open(io.BytesIO(img_data))
                    text = pytesseract.image_to_string(img)
                    logger.debug(f"Used OCR for page {page_num + 1}")
                except Exception as e:
                    logger.warning(f"OCR failed: {str(e)}")

            full_text.append(text)

        return '\n'.join(full_text)

    def _extract_toc_from_first_100_pages(self) -> List[Tuple[str, str, str]]:
        """
        Extract TOC from first 100 pages using cached full_text
        FIXED: Find the TOC that contains actual section numbers (not bid document TOC)
        Returns: List of (display_format, search_format, section_name)
        Example: [("01 33 00", "013300", "SUBMITTAL PROCEDURES"), ...]
        """
        toc_sections = []

        # Calculate text range - search up to 150 pages for TOC
        max_pages = min(150, len(self.doc))
        max_chars = min(len(self.full_text), max_pages * 3000)

        logger.info(f"Searching for TABLE OF CONTENTS in first {max_pages} pages (~{max_chars} chars)")

        search_text = self.full_text[:max_chars]

        # Find ALL occurrences of TABLE OF CONTENTS
        toc_pattern = r'TABLE\s+OF\s+CONTENTS'
        all_toc_matches = list(re.finditer(toc_pattern, search_text, re.IGNORECASE))

        if not all_toc_matches:
            logger.warning("TABLE OF CONTENTS not found in first 100 pages")
            return []

        logger.info(f"Found {len(all_toc_matches)} occurrence(s) of TABLE OF CONTENTS")

        # Try each TOC to find the one with section numbers
        # Pattern: optional whitespace + 5-6 digits (or 2+2+2 with spaces)
        section_num_pattern = re.compile(r'^\s*(\d{2}\s+\d{2}\s+\d{2}|\d{5,6})', re.MULTILINE)

        best_toc_match = None
        best_section_count = 0

        for i, toc_match in enumerate(all_toc_matches):
            toc_start = toc_match.start()

            # Get a window of text after this TOC
            window_size = 50000
            toc_end = min(toc_start + window_size, len(search_text))

            # Check for section boundary
            section_match = self.section_header_pattern.search(search_text, toc_start + 100)
            if section_match:
                toc_end = min(section_match.start(), toc_end)

            toc_text = search_text[toc_start:toc_end]

            # Count how many section numbers this TOC has
            section_numbers = section_num_pattern.findall(toc_text)
            section_count = len(section_numbers)

            logger.info(f"  TOC #{i+1} at position {toc_start}: {section_count} section numbers")

            if section_count > best_section_count:
                best_section_count = section_count
                best_toc_match = toc_match

        if not best_toc_match:
            logger.warning("No TOC found with section numbers")
            return []

        # Use the best TOC (the one with most section numbers)
        toc_start = best_toc_match.start()
        logger.info(f"Using TOC at position {toc_start} with {best_section_count} section numbers")

        # Window size: 50,000 chars
        window_size = 50000
        toc_end = min(toc_start + window_size, len(search_text))

        # Check for section boundary
        section_match = self.section_header_pattern.search(search_text, toc_start + 100)
        if section_match:
            toc_end = min(section_match.start(), toc_end)

        toc_text = search_text[toc_start:toc_end]
        lines = toc_text.split('\n')

        logger.info(f"Parsing TOC ({len(lines)} lines)")

        # Extract section numbers
        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Use pre-compiled pattern
            match = self.toc_line_pattern.match(line)
            if match:
                section_num_raw = match.group(1).strip()
                section_name = match.group(2).strip()

                # Store both: display format (with spaces) and search format (without spaces)
                # If section has spaces, keep them for display
                if ' ' in section_num_raw:
                    section_num_display = section_num_raw  # Keep "01 33 00"
                    section_num_search = section_num_raw.replace(' ', '')  # "013300" for searching
                else:
                    section_num_display = section_num_raw
                    section_num_search = section_num_raw

                # Keep section number as-is
                section_name = self._clean_section_name(section_name)

                # Avoid duplicates
                if not any(sec[0] == section_num_display for sec in toc_sections):
                    # Store (display_format, search_format, name)
                    toc_sections.append((section_num_display, section_num_search, section_name))
                    logger.debug(f"TOC entry: {section_num_display} - {section_name}")

        logger.info(f"Found {len(toc_sections)} sections in TOC")
        return toc_sections

    def _scan_pdf_for_sections(self) -> list:
        """
        Fallback method: Scan PDF directly for SECTION headers when no TOC available
        Returns list of (section_num_display, section_num_search, section_name) tuples
        """
        logger.info("Scanning full PDF for SECTION headers...")
        sections = []

        # Pattern to match SECTION headers
        # Matches: SECTION 01 33 00 or SECTION 013300
        section_pattern = re.compile(
            r'(?:^|\n)\s*SECTION\s+((?:\d{2}\s+\d{2}\s+\d{2}|\d{5,6})(?:\.\d+)?)\s*[-–]?\s*(.+?)(?=\n|$)',
            re.MULTILINE | re.IGNORECASE
        )

        # Find all section headers in full PDF text
        matches = section_pattern.finditer(self.full_text)

        for match in matches:
            section_num_raw = match.group(1).strip()
            section_name = match.group(2).strip()

            # Clean section name (remove page numbers, etc.)
            section_name = self._clean_section_name(section_name)

            # Normalize section number
            if ' ' in section_num_raw:
                # Format: "01 33 00"
                section_num_display = section_num_raw
                section_num_search = section_num_raw.replace(' ', '')
            else:
                # Format: "013300"
                section_num_search = section_num_raw
                # Convert to display format with spaces: "01 33 00"
                if len(section_num_search) == 6:
                    section_num_display = f"{section_num_search[0:2]} {section_num_search[2:4]} {section_num_search[4:6]}"
                else:
                    section_num_display = section_num_search

            # Avoid duplicates
            if not any(sec[0] == section_num_display for sec in sections):
                sections.append((section_num_display, section_num_search, section_name))
                logger.debug(f"Found section: {section_num_display} - {section_name}")

        logger.info(f"Scanned PDF and found {len(sections)} sections")
        return sections

    def _clean_section_name(self, name: str) -> str:
        """Clean section name"""
        name = name.strip()
        name = re.sub(r'\s+\d+\s*$', '', name)  # Remove page numbers
        name = ' '.join(name.split())  # Normalize whitespace
        name = name.upper()
        return name

    def _process_section(self, section_num_display: str, section_num_search: str, section_name: str):
        """
        Process a single section
        Args:
            section_num_display: Section number for display (e.g., "01 33 00")
            section_num_search: Section number for searching (e.g., "013300")
            section_name: Section name
        """
        logger.debug(f"Processing section: {section_num_display} - {section_name}")

        # Check if this is a SUBMITTAL section (extract fully)
        is_submittal_section = "SUBMITTAL" in section_name.upper()

        # Find section text in cached full_text using search format
        section_text = self._extract_section_text(section_num_search)
        if not section_text:
            logger.warning(f"Could not find section {section_num_display} in PDF")
            return

        if is_submittal_section:
            # Extract entire section - use display format for output
            logger.debug(f"Extracting entire section {section_num_display} (SUBMITTAL section)")
            rows = self._extract_full_section(section_num_display, section_name, section_text)
            if rows:
                self.sections_data.extend(rows)
                self.sections_with_submittals.append((section_num_display, section_name))
        else:
            # Extract only paragraphs under submittal subsections - use display format
            rows = self._extract_submittal_subsection(section_num_display, section_name, section_text)
            if rows:
                self.sections_data.extend(rows)
                self.sections_with_submittals.append((section_num_display, section_name))

    def _extract_section_text(self, section_num: str) -> Optional[str]:
        """
        Extract section text from cached full_text using boundaries
        FIXED: Only match section headers at line start (with optional leading whitespace)
        FIXED: Only match "SECTION" in uppercase to avoid false matches with "Section" in text
        FIXED: Handle section numbers with spaces (e.g., "01 11 00") and without (e.g., "011100")
        """
        # Try both formats:
        # 1. Without spaces: "SECTION 011100"
        # 2. With spaces: "SECTION 01 11 00" (if section_num is 6 digits)

        patterns_to_try = [section_num]

        # If section number is 6 digits, also try with spaces
        if len(section_num) == 6 and section_num.isdigit():
            # "011100" → "01 11 00"
            spaced = f"{section_num[0:2]} {section_num[2:4]} {section_num[4:6]}"
            patterns_to_try.append(spaced)

        match = None
        matched_pattern = None

        for pattern_num in patterns_to_try:
            # Find section start - must be at beginning of line (with optional spaces)
            # Pattern: start of string OR newline, then optional spaces, then SECTION (uppercase)
            section_pattern = rf'(?:^|\n)\s*SECTION\s+{re.escape(pattern_num)}(?:\s|$|\n)'
            match = re.search(section_pattern, self.full_text, re.MULTILINE)  # NO re.IGNORECASE!

            if match:
                matched_pattern = pattern_num
                logger.debug(f"Found section {section_num} using pattern '{pattern_num}'")
                break

        if not match:
            return None

        # Find the actual start of "SECTION" text (skip leading newline/spaces)
        start_pos = match.start()
        while start_pos < len(self.full_text) and self.full_text[start_pos] in '\n ':
            start_pos += 1

        # Find next section or end
        # Look for next SECTION at line start (uppercase only)
        # Try both formats: 5 or 6 digits
        next_section_pattern = r'(?:^|\n)\s*SECTION\s+\d{5,6}'  # Without spaces: 5-6 digits
        next_section_match = re.search(next_section_pattern, self.full_text[start_pos + 1:], re.MULTILINE)

        # Also try with spaces
        next_section_pattern_spaced = r'(?:^|\n)\s*SECTION\s+\d{2}\s+\d{2}\s+\d{2}'
        next_section_match_spaced = re.search(next_section_pattern_spaced, self.full_text[start_pos + 1:], re.MULTILINE)

        # Use whichever is found first
        if next_section_match and next_section_match_spaced:
            if next_section_match.start() < next_section_match_spaced.start():
                next_section_match = next_section_match
            else:
                next_section_match = next_section_match_spaced
        elif next_section_match_spaced:
            next_section_match = next_section_match_spaced

        if next_section_match:
            end_pos = start_pos + 1 + next_section_match.start()
            # Skip leading whitespace of next section
            while end_pos > start_pos and self.full_text[end_pos - 1] in '\n ':
                end_pos -= 1
        else:
            end_pos = len(self.full_text)

        return self.full_text[start_pos:end_pos]

    def _merge_multiline_subsections(self, lines: List[str]) -> List[str]:
        """
        Merge multi-line subsection headings AND hierarchy items
        Examples:
        - "1.04" + "SUBMITTALS" → "1.04 SUBMITTALS"
        - "A." + "The Contractor shall..." → "A. The Contractor shall..."
        """
        merged_lines = []
        i = 0

        # Patterns for items that might be on their own line
        subsection_num_pattern = re.compile(r'^(\d+\.\d+)$')  # Just "1.04"
        single_letter_pattern = re.compile(r'^([A-Z])\.$')     # Just "A."
        single_lowercase_pattern = re.compile(r'^([a-z])\.$')  # Just "a."

        while i < len(lines):
            line = lines[i].strip()

            # Check if current line is just a subsection number (e.g., "1.04")
            if subsection_num_pattern.match(line) and i + 1 < len(lines):
                next_line = lines[i + 1].strip()

                if next_line and not subsection_num_pattern.match(next_line):
                    # Merge the lines
                    merged = f"{line} {next_line}"
                    merged_lines.append(merged)
                    logger.debug(f"Merged subsection: '{line}' + '{next_line}' → '{merged}'")
                    i += 2  # Skip both lines
                    continue

            # Check if current line is just "A." or "B." etc.
            elif single_letter_pattern.match(line) and i + 1 < len(lines):
                next_line = lines[i + 1].strip()

                if next_line and len(next_line) > 0:
                    # Merge the lines
                    merged = f"{line} {next_line}"
                    merged_lines.append(merged)
                    logger.debug(f"Merged letter: '{line}' + '{next_line}' → '{merged}'")
                    i += 2  # Skip both lines
                    continue

            # Check if current line is just "a." or "b." etc.
            elif single_lowercase_pattern.match(line) and i + 1 < len(lines):
                next_line = lines[i + 1].strip()

                if next_line and len(next_line) > 0:
                    # Merge the lines
                    merged = f"{line} {next_line}"
                    merged_lines.append(merged)
                    logger.debug(f"Merged lowercase: '{line}' + '{next_line}' → '{merged}'")
                    i += 2  # Skip both lines
                    continue

            merged_lines.append(line)
            i += 1

        return merged_lines

    def _merge_continuation_lines(self, lines: List[str]) -> List[Dict]:
        """
        FIXED: Merge continuation lines with their parent hierarchy item
        FIXED: Handle multi-line hierarchy items (e.g., "A." on one line, text on next)

        Returns list of dicts with 'level' and 'text' keys
        """
        # FIRST: Merge multi-line hierarchy items
        single_letter_pattern = re.compile(r'^([A-Z])\.\s*$')
        single_lowercase_pattern = re.compile(r'^([a-z])\.\s*$')
        single_number_pattern = re.compile(r'^(\d{1,2})\.\s*$')

        preprocessed_lines = []
        i = 0
        while i < len(lines):
            line = lines[i].strip()

            if not line:  # Skip empty lines
                i += 1
                continue

            # Check if line is just a hierarchy marker
            if (single_letter_pattern.match(line) or
                single_lowercase_pattern.match(line) or
                single_number_pattern.match(line)) and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line:
                    # Merge
                    merged = f"{line} {next_line}"
                    preprocessed_lines.append(merged)
                    logger.debug(f"Merged hierarchy item: '{line}' + '{next_line[:50]}'")
                    i += 2
                    continue

            preprocessed_lines.append(line)
            i += 1

        # NOW: Process the merged lines
        merged = []
        current_item = None
        current_level = None

        for line in preprocessed_lines:
            # REMOVED length check here - hierarchy items can be short
            if not line:
                continue

            # Clean the line
            cleaned = self._clean_line(line)
            if not cleaned:
                continue

            # Check if this line matches a hierarchy pattern
            level_num = self._detect_hierarchy(cleaned)

            if level_num:
                # This is a new hierarchy item
                # Save previous item if exists
                if current_item:
                    merged.append({
                        'level': current_level,
                        'text': current_item
                    })

                # Start new item
                current_item = cleaned
                current_level = level_num
            else:
                # This is a continuation line
                if current_item:
                    # Append to current item with space
                    current_item += ' ' + cleaned
                else:
                    # No parent, treat as Level 10
                    merged.append({
                        'level': 10,
                        'text': cleaned
                    })

        # Don't forget the last item
        if current_item:
            merged.append({
                'level': current_level,
                'text': current_item
            })

        return merged

    def _extract_full_section(self, section_num: str, section_name: str,
                              section_text: str) -> List[Dict]:
        """
        Extract entire section with all content
        For sections with "SUBMITTAL" in the name
        """
        rows = []
        lines = section_text.split('\n')

        # FIXED: Merge multi-line subsection headings
        lines = self._merge_multiline_subsections(lines)

        # First row with section info
        first_row = {
            'Section Number': section_num,
            'Section Name': section_name,
        }
        for i in range(1, 11):
            first_row[f'Level {i}'] = ''
        rows.append(first_row)

        # Process lines - find section start
        section_started = False
        section_lines = []

        for line in lines:
            if not line:
                continue

            # Check if section header
            if self.section_header_pattern.search(line):
                section_started = True
                continue

            if not section_started:
                continue

            # Skip section name line
            if line.upper().strip() == section_name.upper():
                continue

            section_lines.append(line)

        # FIXED: Merge continuation lines
        merged_items = self._merge_continuation_lines(section_lines)

        # Create rows from merged items
        for item in merged_items:
            row = {
                'Section Number': '',
                'Section Name': '',
            }
            for i in range(1, 11):
                row[f'Level {i}'] = ''

            row[f'Level {item["level"]}'] = item['text']
            rows.append(row)

        logger.debug(f"Extracted {len(rows)} rows from full section {section_num}")
        return rows

    def _extract_submittal_subsection(self, section_num: str, section_name: str,
                                      section_text: str) -> List[Dict]:
        """
        FIXED: Extract only paragraphs under submittal subsections until next subsection
        For sections WITHOUT "SUBMITTAL" in the name
        """
        rows = []
        lines = section_text.split('\n')

        # FIXED: Merge multi-line subsection headings
        lines = self._merge_multiline_subsections(lines)

        # FIXED: Check if there's a Level 2 subsection with SUBMITTAL keyword (UPPERCASE only)
        # Must match pattern like "1.03 SUBMITTALS" or "1.3 - SUBMITTALS" or "1.3 FORM OF SUBMITTALS"
        has_submittal_subsection = False
        for line in lines:
            level_2_match = self.hierarchy_patterns[1].match(line)  # Level 2 pattern
            if level_2_match:
                subsection_num = level_2_match.group(1)
                subsection_text = level_2_match.group(2)

                # CRITICAL FIX: Check if "SUBMITTAL" appears in UPPERCASE in the subsection text
                # This filters out lowercase "submittals" in regular text like "reviewing submittals"
                if 'SUBMITTAL' in subsection_text:  # Must be uppercase
                    has_submittal_subsection = True
                    logger.debug(f"Found SUBMITTAL subsection in {section_num}: {subsection_num} {subsection_text}")
                    break

        if not has_submittal_subsection:
            return []

        # First row
        first_row = {
            'Section Number': section_num,
            'Section Name': section_name,
        }
        for i in range(1, 11):
            first_row[f'Level {i}'] = ''
        rows.append(first_row)

        # NEW APPROACH: Scan entire section and extract ALL submittal subsections
        # even if they're not consecutive
        in_submittal_subsection = False
        subsection_lines = []

        for line in lines:
            if not line or len(line) < 3:
                continue

            # Check if this line is a subsection heading (Level 2)
            level_2_match = self.hierarchy_patterns[1].match(line)

            if level_2_match:
                subsection_num = level_2_match.group(1)
                subsection_text = level_2_match.group(2)

                # Check if this is a SUBMITTAL subsection
                is_submittal_subsection = 'SUBMITTAL' in subsection_text

                if in_submittal_subsection:
                    # We were in a submittal subsection
                    # Process the previous one before starting new subsection
                    logger.debug(f"Processing previous submittal subsection in {section_num}")
                    if subsection_lines:
                        merged_items = self._merge_continuation_lines(subsection_lines)
                        for item in merged_items:
                            row = {
                                'Section Number': '',
                                'Section Name': '',
                            }
                            for i in range(1, 11):
                                row[f'Level {i}'] = ''
                            if item['level'] != 2:
                                row[f'Level {item["level"]}'] = item['text']
                                rows.append(row)

                    # Clear lines for next subsection
                    subsection_lines = []

                # Now handle the new subsection
                if is_submittal_subsection:
                    # This is a submittal subsection - start extracting
                    in_submittal_subsection = True
                    logger.debug(f"Entering submittal subsection in {section_num}: {line}")

                    # Add the subsection heading as Level 2
                    row = {
                        'Section Number': '',
                        'Section Name': '',
                    }
                    for i in range(1, 11):
                        row[f'Level {i}'] = ''
                    row['Level 2'] = line
                    rows.append(row)
                    continue
                else:
                    # Non-submittal subsection - skip it but keep scanning
                    in_submittal_subsection = False
                    logger.debug(f"Skipping non-submittal subsection in {section_num}: {line}")
                    continue

            # Collect lines if we're inside a submittal subsection
            if in_submittal_subsection:
                subsection_lines.append(line)

        # Process last submittal subsection if any
        if subsection_lines:
            logger.debug(f"Processing last submittal subsection in {section_num}")
            merged_items = self._merge_continuation_lines(subsection_lines)

            # Create rows from merged items
            for item in merged_items:
                row = {
                    'Section Number': '',
                    'Section Name': '',
                }
                for i in range(1, 11):
                    row[f'Level {i}'] = ''

                # Skip Level 2 (subsection headings) as they're already added
                if item['level'] != 2:
                    row[f'Level {item["level"]}'] = item['text']
                    rows.append(row)

        logger.debug(f"Extracted {len(rows)} rows from submittal subsection in {section_num}")
        return rows

    def _clean_line(self, line: str) -> str:
        """Clean line by removing footers and artifacts"""
        for pattern in self.footer_patterns:
            if pattern.search(line):
                return ""

        # Skip common artifacts
        if re.match(r'^(PAGE|SECTION NUMBER|PROJECT MANUAL)$', line.upper()):
            return ""

        # Skip if line is just a section number
        if re.match(r'^\d{5}$', line):
            return ""

        # Skip project names and locations
        if 'Jacksonville' in line or 'Robena Road' in line:
            return ""

        return line

    def _detect_hierarchy(self, line: str) -> Optional[int]:
        """
        Detect hierarchy level
        Returns level number (1-10) or None
        """
        for level, pattern in enumerate(self.hierarchy_patterns, 1):
            if pattern.match(line):
                return level
        return None

    def _format_sections_output(self) -> openpyxl.Workbook:
        """
        Format sections data as Excel workbook with multiple sheets
        FIXED: Put SUBMITTAL sections first, then alphabetical
        """
        wb = openpyxl.Workbook()
        default_sheet = wb.active

        # Group sections by number
        # FIXED: Only group rows that actually belong to each section
        sections_by_number = {}

        # First, identify where each section starts
        section_starts = []
        for idx, row in enumerate(self.sections_data):
            section_num = row.get('Section Number', '').strip()
            if section_num:
                section_starts.append((idx, section_num))

        # Add end marker
        section_starts.append((len(self.sections_data), None))

        # Group rows by section
        for i in range(len(section_starts) - 1):
            start_idx, section_num = section_starts[i]
            end_idx, _ = section_starts[i + 1]

            # All rows from start_idx to end_idx belong to this section
            section_rows = self.sections_data[start_idx:end_idx]
            sections_by_number[section_num] = section_rows

        logger.info(f"Grouped data into {len(sections_by_number)} sections")

        if not self.sections_with_submittals:
            logger.warning("No sections with submittals found")
            default_sheet.title = "No Data"
            default_sheet.cell(1, 1, "No sections with submittals were found in the PDF")
            return wb

        # FIXED: Sort sections - SUBMITTAL sections first, then alphabetical
        submittal_sections = []
        other_sections = []

        for section_num, section_name in self.sections_with_submittals:
            if "SUBMITTAL" in section_name.upper():
                submittal_sections.append((section_num, section_name))
            else:
                other_sections.append((section_num, section_name))

        # Sort each group alphabetically
        submittal_sections.sort(key=lambda x: x[0])
        other_sections.sort(key=lambda x: x[0])

        # Combine: submittal sections first, then others
        sorted_sections = submittal_sections + other_sections

        logger.info(f"Sheet order: {len(submittal_sections)} SUBMITTAL sections first, then {len(other_sections)} others")

        # Create sheets in sorted order
        first_sheet_created = False
        for section_num, section_name in sorted_sections:
            sheet_name = f"{section_num}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)

            if not first_sheet_created:
                wb.remove(default_sheet)
                first_sheet_created = True

            # Headers
            headers = ['Section Number', 'Section Name',
                      'Level 1', 'Level 2', 'Level 3', 'Level 4', 'Level 5',
                      'Level 6', 'Level 7', 'Level 8', 'Level 9', 'Level 10']

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(1, col_idx, header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Data
            section_rows = sections_by_number.get(section_num, [])

            if not section_rows:
                logger.warning(f"No data found for section {section_num}")
                continue

            for row_idx, data_row in enumerate(section_rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = data_row.get(header, '')
                    if value:
                        ws.cell(row_idx, col_idx, value)

            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            logger.debug(f"Created sheet '{sheet_name}' with {len(section_rows)} rows")

        logger.info(f"Created workbook with {len(wb.sheetnames)} sheets")
        return wb

    def _format_log_output(self, template_path: str = None) -> openpyxl.Workbook:
        """
        Format log using template
        """
        if template_path and os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb['Submittal Resp']
            logger.info(f"Using template from: {template_path}")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Submittal Resp'
            headers = ['Section', '', '', '', 'Description', '', 'Type', 'Sub/Vendor',
                      'Due', "Recv'd", 'Float', 'Submit', 'HIDE', 'Return', 'Status']
            for col_idx, header in enumerate(headers, 1):
                ws.cell(6, col_idx, header)
            logger.info("No template found, creating basic structure")

        # Populate data starting from row 7
        start_row = 7
        for idx, (section_num, section_name) in enumerate(self.sections_with_submittals):
            row_num = start_row + idx
            ws.cell(row_num, 1, section_num)
            ws.cell(row_num, 5, f"{section_num} - {section_name}")

        logger.info(f"Populated {len(self.sections_with_submittals)} sections in log")
        return wb


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1:
        pdf_file = sys.argv[1]

        print(f"Testing extractor on: {pdf_file}")

        import time
        start = time.time()

        extractor = SubmittalExtractor(pdf_file)
        result = extractor.extract()

        # Save output
        result['sections'].save("test_sections.xlsx")
        result['log'].save("test_log.xlsx")

        elapsed = time.time() - start

        print(f"\n=== EXTRACTION RESULTS ===")
        print(f"Time: {elapsed:.2f} seconds")
        print(f"Sections with submittals: {len(extractor.sections_with_submittals)}")
        print(f"Files saved: test_sections.xlsx, test_log.xlsx")
    else:
        print("Usage: python extractor.py <path_to_pdf>")